
import streamlit as st
import pandas as pd
import os
from pathlib import Path
import openpyxl
from datetime import date
import numpy as np
from streamlit_extras.switch_page_button import switch_page
from datetime import timedelta

st.set_page_config(page_title='Science Matters')
st.title('Invoice')
##################retrieve info from excel ***************************

# Path Setting:
try:
    current_path = Path(__file__).parent.absolute()  # Get the file path for this py file
except:
    current_path = (Path.cwd())

# import the inv_sheet from the excel file

filepath = os.path.join(current_path, 'ERP.xlsx')
wb = openpyxl.load_workbook(filepath)
#df = pd.read_excel(filepath, engine='openpyxl')
psheet = wb['Parent']
hsheet = wb['Holidays']
psheet.delete_rows(1)
hsheet.delete_rows(1)
#put holidays dates into a numpy array
hlist3 = []
hlist4 = []
hlist5 = []
hlist6 = []
for row in hsheet.iter_rows():
    hlist3.append(np.datetime64(row[0].value))
    hlist4.append(np.datetime64(row[1].value))
    hlist5.append(np.datetime64(row[2].value))
    hlist6.append(np.datetime64(row[3].value))

HP3 = np.array(hlist3, dtype='datetime64[D]')
HP4 = np.array(hlist4, dtype='datetime64[D]')
HP5 = np.array(hlist5, dtype='datetime64[D]')
HP6 = np.array(hlist6, dtype='datetime64[D]')


#put worksheet into dataframe
price_df = pd.read_excel(filepath, engine='openpyxl', sheet_name='Price')
parent_df = pd.read_excel(filepath, engine='openpyxl', sheet_name='Parent')
# set values for account for easy search
l_sname = parent_df['stud_name'].tolist()
options = parent_df['acct_no'].tolist()
l_sname.insert(0, "All Account")
options.insert(0, "all")
dic = dict(zip(options, l_sname))




df1A = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
df1B = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
df2A = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
df2B = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
df3A = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
df3B = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
df4A = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
df4B = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
df5A = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
df5B = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])

###############set all session state variables###############
if 'dtfrm' not in st.session_state:
    st.session_state.dtfrm = date.today()
if 'dtto' not in st.session_state:
    st.session_state.dtto = date.today()
if 'acct' not in st.session_state:
    st.session_state.acct = ""
if 'bill' not in st.session_state:
    st.session_state['bill'] = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])
if 'duedt' not in st.session_state:
    st.session_state.duedt = date.today()
#####################define methods#####################
def cal_qty(zclass, zlevel):
    zday = int(zclass[0])
    if zday == 1:
        day_txt = 'Mon'
    if zday == 2:
        day_txt = 'Tue'
    if zday == 3:
        day_txt = 'Wed'
    if zday == 4:
        day_txt = 'Thu'
    if zday == 5:
        day_txt = 'Fri'
    zlevel = 'H' + zlevel
    zqty = np.busday_count(st.session_state['dtfrm'], st.session_state['dtto'] + timedelta(days=1), weekmask=day_txt, holidays=globals()[zlevel])
    return zqty

def add_class_billing(s_class):
    #df_ret = pd.DataFrame(columns=['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty', 'Price', 'Date_From', 'Date_To', 'Due_Date'])
    for row in psheet.iter_rows():
        if (st.session_state.acct == row[0].value or st.session_state.acct == 'all') and (s_class == row[6].value or s_class ==""):
            level = row[5].value
            no_class = cal_qty(row[6].value, level)
            if no_class > 0:
                unit_pr = price_df.loc[price_df['mcode'] == row[5].value, 'amount'].iloc[0]
                add_row = pd.DataFrame({'Account': [row[0].value], 'Parent': [row[1].value], 'Student': [row[2].value], 'Class': [row[6].value], 'Item': [row[5].value], 'Qty': [no_class], 'Price': [unit_pr], 'Date_From': [st.session_state.dtfrm], 'Date_To': [st.session_state.dtto], 'Due_Date': [st.session_state.duedt]})
                st.session_state.bill = pd.concat([st.session_state.bill, add_row], ignore_index = True)

def add_itm_billing(s_acct, s_itm, s_qty, s_price):
    for row in psheet.iter_rows():
        if (s_acct == row[0].value or s_acct == 'all'):
            add_row = pd.DataFrame({'Account': [row[0].value], 'Parent': [row[1].value], 'Student': [row[2].value], 'Class': [row[6].value], 'Item': s_itm, 'Qty': s_qty, 'Price': s_price, 'Date_From': [st.session_state.dtfrm], 'Date_To': [st.session_state.dtto], 'Due_Date': [st.session_state.duedt]})
            st.session_state.bill = pd.concat([st.session_state.bill, add_row], ignore_index = True)

def add_one_time_billing(xname, xsname, xitm, xqty, xprice):
    add_row = pd.DataFrame({'Account': 199999, 'Parent': xname, 'Student': xsname, 'Class': "None", 'Item': xitm, 'Qty': xqty, 'Price': xprice, 'Date_From': "None", 'Date_To': "None", 'Due_Date': [st.session_state.duedt]})
    st.session_state.bill = pd.concat([st.session_state.bill, add_row], ignore_index = True)

###################set parameters for bill***********************
with st.form(key = "period"):
    c1, c2, c3 = st.columns(3)
    with c1:
        inv_dtfrm = st.date_input("Invoice Date From")
        p1_acct = st.selectbox("Account", options , format_func=lambda x: dic[x])
    with c2:
        inv_dtto = st.date_input("Invoice Date To")
        p1_class = st.text_input("Class. Use \',\' for more than one")

    with c3:
        inv_duedt = st.date_input("Invoice Due Date")

    if st.form_submit_button(label= 'Add billing for classes based on period and account'):
        st.session_state.dtfrm = inv_dtfrm
        st.session_state.dtto = inv_dtto
        st.session_state.acct = p1_acct
        st.session_state.duedt = inv_duedt
        if p1_class != "":
            l_class = p1_class.split(",")
            for x in l_class:
                add_class_billing(x)
        else:

            add_class_billing("")
        #st.experimental_rerun()
        st.write("Billing for Classes added!")

#st.text("Invoice Date from : " + str(st.session_state.dtfrm) + " to " + str(st.session_state.dtto))
st.subheader('Add One Time billing here')
with st.form(key = "period2"):
    st.text("<-----For one time billing----->")
    c1, c2, c3 = st.columns(3)
    with c1:
        inv_duedt = st.date_input("Invoice Due Date")
        inv_itm = st.text_input("Item Description")
    with c2:
        inv_name = st.text_input(" Parent Name")
        inv_qty = st.number_input("Quantity")
    with c3:
        s_name = st.text_input("Student Name")
        inv_price = st.number_input("Price")

    if st.form_submit_button(label= 'Add one time billing'):
        st.session_state.dtfrm = ""
        st.session_state.dtto = ""
        st.session_state.acct = "None"
        st.session_state.duedt = inv_duedt
        st.session_state.acct = "199999"
        add_one_time_billing(inv_name, s_name, inv_itm, inv_qty, inv_price)
        st.experimental_rerun()

# display all billing items
st.session_state.bill.sort_values(by=['Class', 'Account'], inplace=True)
st.dataframe(st.session_state.bill)

st.subheader('Add Additional Invoice items here')
with st.expander('Add Invoice Items'):
    with st.form(key="add_items"):
        c1, c2 = st.columns(2)
        with c1:
        #to add items
            p_acct = st.selectbox("Account", options , format_func=lambda x: dic[x])
            p_qty = st.text_input("Quantity")
        with c2:
            p_item = st.selectbox("Item", ('Add. Class', 'Free Class', 'Material Cost', 'Deposit', 'Discount', 'Workshop', 'Others'))
            p_price = st.number_input("Price")

        if st.form_submit_button(label='Add item'):
            #l_acct = p_acct.split(",")
            #for x in l_acct:
            #    x = int(x)
            #    c_parent = st.session_state.bill.loc[st.session_state.bill['Account'] == x, 'Parent'].iloc[0]
            #    c_student = st.session_state.bill.loc[st.session_state.bill['Account'] == x, 'Student'].iloc[0]
            #    c_class = st.session_state.bill.loc[st.session_state.bill['Account'] == x, 'Class'].iloc[0]
            #    add_row = pd.DataFrame({'Account': [x], 'Parent': [c_parent], 'Student': [c_student], 'Class': [c_class], 'Item': [p_item], 'Qty': [p_qty], 'Price': ["{:.2f}".format(p_price)], 'Date_From': [st.session_state.dtfrm], 'Date_To': [st.session_state.dtto], 'Due_Date': [st.session_state.duedt]})
            #   st.session_state.bill = pd.concat([st.session_state.bill, add_row], ignore_index = True)
            add_itm_billing(p_acct, p_item, p_qty, p_price)
            st.experimental_rerun()

st.subheader('Remove Invoice items here')
with st.expander('Remove Invoice Items'):
    p_index = st.text_input("Index no")
    #p_class = st.selectbox("Class", ('1A', '1B', '2A', '2B', '3A', '3B', '4A', '4B', '5A', '5B'))
    c1, c2 = st.columns(2)
    with c1:
        if st.button('Remove item'):
            p_index = int(p_index)
            st.session_state.bill.drop(p_index, axis=0, inplace=True)
            st.experimental_rerun()
    with c2:
        if st.button('CLEAR ALL ITEMS!'):
            del st.session_state.bill
            st.experimental_rerun()

with st.expander('Class 1A'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '1A']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)

with st.expander('Class 1B'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '1B']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)

with st.expander('Class 2A'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '2A']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)

with st.expander('Class 2B'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '2B']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)

with st.expander('Class 3A'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '3A']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)

with st.expander('Class 3B'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '3B']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)

with st.expander('Class 4A'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '4A']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)

with st.expander('Class 4B'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '4B']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)

with st.expander('Class 5A'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '5A']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)

with st.expander('Class 5B'):
    df1A = st.session_state.bill.loc[st.session_state.bill['Class'] == '5B']
    df1A.sort_values(by=['Account'], inplace=True)
    st.dataframe(df1A)



if st.button('Create Invoice for list'):
    switch_page("Invoice")





import streamlit as st
import pandas as pd
import os
from pathlib import Path
import openpyxl
from datetime import date
#####################pdf imports###################
from borb.pdf import Document
from borb.pdf.page.page import Page
from borb.pdf.canvas.layout.page_layout.multi_column_layout import SingleColumnLayout
from decimal import Decimal
from borb.pdf.canvas.layout.image.image import Image
from borb.pdf.canvas.layout.table.fixed_column_width_table import FixedColumnWidthTable as Table
from borb.pdf.canvas.layout.text.paragraph import Paragraph
from borb.pdf.canvas.layout.layout_element import Alignment
from borb.pdf.pdf import PDF
from borb.pdf.canvas.color.color import HexColor, X11Color
from borb.pdf.canvas.layout.table.fixed_column_width_table import FixedColumnWidthTable as Table
from borb.pdf.canvas.layout.table.table import TableCell
from borb.pdf.canvas.layout.image.barcode import Barcode, BarcodeType
from borb.pdf.canvas.layout.layout_element import LayoutElement
from borb.pdf.canvas.line_art.line_art_factory import LineArtFactory
from borb.pdf.canvas.geometry.rectangle import Rectangle
from borb.pdf.canvas.layout.shape.shape import Shape
from borb.pdf.page.page_size import PageSize
import typing
import random

############################paynowQR#######################
from pyPayNowSg import PayNowConfigFactory, PayNowSerializer


st.set_page_config(page_title='Science Matters')
st.title('Payment Receipt for Invoice/Deposit')
################methods###############################
def add_artwork_upper_right_corner(page: Page) -> None:
    ps: typing.Tuple[Decimal, Decimal] = PageSize.A4_PORTRAIT.value
    N: int = 4
    M: Decimal = Decimal(32)
    x: Decimal = ps[0] - N * 105
    y: Decimal = ps[1] - (0 + 2) * M
    IMAGE_PATH = Path(r"C:\Users\Choon Yong Chong\OneDrive\Documents\Business\Tuition center\SM logo.png")
    Image(image=IMAGE_PATH, width=239, height=27).layout(page, Rectangle(x, y, M, M))
    #page_layout.add(Image(image=IMAGE_PATH, width=100, height=50, margin_top=0, margin_right=0))  # change the size as you wish


def add_gray_artwork_upper_right_corner(page: Page) -> None:
    """
    This method will add a gray artwork of squares and triangles in the upper right corner
    of the given Page
    """
    grays: typing.List[HexColor] = [
        HexColor("A9A9A9"),
        HexColor("D3D3D3"),
        HexColor("DCDCDC"),
        HexColor("E0E0E0"),
        HexColor("E8E8E8"),
        HexColor("F0F0F0"),
    ]
    ps: typing.Tuple[Decimal, Decimal] = PageSize.A4_PORTRAIT.value
    N: int = 4
    M: Decimal = Decimal(32)

    # Draw triangles
    for i in range(0, N):
        x: Decimal = ps[0] - N * M + i * M
        y: Decimal = ps[1] - (i + 1) * M
        rg: HexColor = random.choice(grays)
        Shape(
            points=[(x + M, y), (x + M, y + M), (x, y + M)],
            stroke_color=rg,
            fill_color=rg,
        ).layout(page, Rectangle(x, y, M, M))

    # Draw squares
    for i in range(0, N - 1):
        for j in range(0, N - 1):
            if j > i:
                continue
            x: Decimal = ps[0] - (N - 1) * M + i * M
            y: Decimal = ps[1] - (j + 1) * M
            rg: HexColor = random.choice(grays)
            Shape(
                points=[(x, y), (x + M, y), (x + M, y + M), (x, y + M)],
                stroke_color=rg,
                fill_color=rg,
            ).layout(page, Rectangle(x, y, M, M))


def add_colored_artwork_bottom_right_corner(page: Page) -> None:
    """
    This method will add a blue/purple artwork of lines
    and squares to the bottom right corner
    of the given Page
    """
    ps: typing.Tuple[Decimal, Decimal] = PageSize.A4_PORTRAIT.value
    # Line
    r: Rectangle = Rectangle(Decimal(0), Decimal(32), ps[0], Decimal(8))
    Shape(
        points=LineArtFactory.rectangle(r),
        stroke_color=HexColor("fbb03f"),
        fill_color=HexColor("fbb03f"),
    ).layout(page, r)

def _build_invoice_information(df):
    dtfrom = df.iloc[0]['Invdt_frm']
    dtto = df.iloc[0]['Invdt_to']
    duedt = df.iloc[0]['Due']
    inv = df.iloc[0]['Invoice']


    table_001 = Table(number_of_rows=5, number_of_columns=6)

    table_001.add(TableCell(Paragraph("Science Matters Pte Ltd", font="Helvetica-Bold"), col_span=2))
    table_001.add(TableCell(Paragraph(" "), col_span=2))
    table_001.add(TableCell(Paragraph("Payment Receipt", font_color=HexColor("#fbb03f"), font_size=Decimal(14),font="Helvetica-Bold", horizontal_alignment=Alignment.CENTERED), col_span=2))

    table_001.add(TableCell(Paragraph("https://www.sciencematters.com.sg", font_size=Decimal(10)), col_span=2))
    table_001.add(TableCell(Paragraph(" "), col_span=4))

    if dtfrom == "None":
        dt_frm_desp = " "
        dtfrom = " "
    else:
        dt_frm_desp = "Bill Date From"
        #dtfrom = ": %d/%d/%d" % (dtfrom.day, dtfrom.month, dtfrom.year)

    if dtto == "None":
        dt_to_desp = " "
        dtto = " "
    else:
        dt_to_desp = "Bill Date From"
        #dtto = ": %d/%d/%d" % (dtto.day, dtto.month, dtto.year)

    table_001.add(Paragraph(dt_frm_desp, font_size=Decimal(10), font="Helvetica-Bold"))
    table_001.add(Paragraph(dtfrom, font_size=Decimal(10)))
    table_001.add(TableCell(Paragraph(" "), col_span=2))
    table_001.add(Paragraph("Date :", font_size=Decimal(10), font="Helvetica-Bold", horizontal_alignment=Alignment.RIGHT))
    now = date.today().strftime('%d/%m/%Y')
    table_001.add(Paragraph(now, font_size=Decimal(10)))


    table_001.add(Paragraph(dt_to_desp, font_size=Decimal(10), font="Helvetica-Bold"))
    table_001.add(Paragraph(dtto, font_size=Decimal(10)))
    table_001.add(TableCell(Paragraph(" "), col_span=2))
    table_001.add(Paragraph("Receipt No :", font_size=Decimal(10), font="Helvetica-Bold", horizontal_alignment=Alignment.RIGHT))
    table_001.add(Paragraph(receipt_no, font_size=Decimal(10)))

    table_001.add(TableCell(Paragraph(" "), col_span=4))
    table_001.add(Paragraph("Due Date :", font_size=Decimal(10), font="Helvetica-Bold", horizontal_alignment=Alignment.RIGHT))
    #table_001.add(Paragraph("%d/%d/%d" % (duedt.day, duedt.month, duedt.year), font_size=Decimal(10)))
    table_001.add(Paragraph(duedt, font_size=Decimal(10)))

    table_001.set_padding_on_all_cells(Decimal(2), Decimal(2), Decimal(2), Decimal(2))
    table_001.no_borders()
    return table_001


def _build_billing_and_shipping_information(m_name):
    table_001 = Table(number_of_rows=3, number_of_columns=1)
    table_001.add(Paragraph("BILL TO",font_size=Decimal(10), font="Helvetica-Bold"))
    table_001.add(Paragraph(m_name,font_size=Decimal(10)))  # BILLING
    table_001.add(Paragraph(" "))
    table_001.set_padding_on_all_cells(Decimal(2), Decimal(2), Decimal(2), Decimal(2))
    table_001.no_borders()
    return table_001


def _build_itemized_description_table(df, m_amt, m_out, m_dt):
    prev_out = df.iloc[0]['Outstanding']

    #count number of items/acct code appears in dataframe
    nrow = 5

    # draw table headers
    table_001 = Table(number_of_rows=nrow, number_of_columns=6)

    table_001.add(TableCell(Paragraph("Description", font_size=Decimal(9)), background_color=HexColor("#D3D3D3"),
                            padding_bottom=Decimal(30),
                            col_span=5))
    table_001.add(TableCell(Paragraph("Amount", font_size=Decimal(9),horizontal_alignment=Alignment.RIGHT),
                                background_color=HexColor("#D3D3D3"),
                                padding_bottom=Decimal(30), ))

    description = 'Payment Received on ' + str(m_dt)
    table_001.add(TableCell(Paragraph(description, font_size=Decimal(9)), col_span=5))
    table_001.add(TableCell(Paragraph("$ " + str(m_amt), font_size=Decimal(9),horizontal_alignment=Alignment.RIGHT)))
    table_001.add(TableCell(Paragraph("_____________________________________________________________________",font_color=HexColor("#D3D3D3"), horizontal_alignment=Alignment.CENTERED),col_span=6))


    table_001.add(TableCell(Paragraph("Previous Balance :", font_size=Decimal(9),
                            horizontal_alignment=Alignment.RIGHT), col_span=5))
    table_001.add(TableCell(Paragraph("$ " + str(prev_out), font_size=Decimal(9),
                                      horizontal_alignment=Alignment.RIGHT)))

    table_001.add(TableCell(Paragraph("Outstanding Amount :",
                            font_size=Decimal(9),
                            font="Helvetica-Bold",
                            horizontal_alignment=Alignment.RIGHT),
                            background_color=HexColor("#fbb03f"),
                            col_span=5))

    table_001.add(TableCell(Paragraph("$ " + str(m_out),
                                      font_size=Decimal(9),
                                      font="Helvetica-Bold",
                                      horizontal_alignment=Alignment.RIGHT),
                            background_color=HexColor("#fbb03f")))

    table_001.set_padding_on_all_cells(Decimal(2), Decimal(2), Decimal(2), Decimal(2))
    table_001.no_borders()
    return table_001

##################retrieve info from excel ***************************

# Path Setting:
try:
    current_path = Path(__file__).parent.absolute()  # Get the file path for this py file
except:
    current_path = (Path.cwd())

filepath = os.path.join(current_path, 'ERP.xlsx')
wb = openpyxl.load_workbook(filepath)
#df = pd.read_excel(filepath, engine='openpyxl')
inv_sheet = wb['Invoice']
last_rno = int(inv_sheet.cell(row=1, column=17).value)

#if 'due' not in st.session_state:
st.session_state['due'] = pd.DataFrame(columns = ['Account', 'Invoice', 'Amt', 'Invdt_frm','Invdt_to','Paid', 'Due', 'Outstanding', 'Name'])
#put all unpaid invoice into dataframe
for row in inv_sheet.iter_rows():
    if row[11].value != 0 and row[11].value != 'outstanding':
        add_row = pd.DataFrame({'Account': [row[0].value], 'Invoice': [row[1].value], 'Amt': [row[3].value], 'Invdt_frm': [row[4].value],'Invdt_to': [row[5].value],'Paid': [row[7].value],'Due': [row[9].value], 'Outstanding': [row[11].value], 'Name': [row[12].value]})
        st.session_state.due = pd.concat([st.session_state.due, add_row], ignore_index=True)

# display all due items
st.session_state.due.sort_values(by=['Account', 'Invoice'], inplace=True)
st.dataframe(st.session_state.due)

###################create form for payment ##################
with st.form(key="payment"):
    p_inv = st.text_input("Invoice number")
    p_amt = st.number_input('Amount. Leave blank if fully paid')
    p_date = st.date_input("Date received")

    # Every form must have a submit button.
    submitted = st.form_submit_button("Submit")
    if submitted:
        df_inv = st.session_state.due.loc[st.session_state.due['Invoice'] == p_inv]
        try:
            inv_amt = float(df_inv.iloc[0]['Amt'])
        except:
            inv_amt = 0
        try:
            p_paid = float(df_inv.iloc[0]['Paid'])
        except:
            p_paid = 0

        if p_amt == 0:
            p_amt = inv_amt
        p_out = inv_amt - p_paid - p_amt
        ttl_paid = p_amt + p_paid
        name = df_inv.iloc[0]['Name']
        receipt_no = p_inv + "_" + str(last_rno)

        # Create document
        pdf = Document()

        # Add page
        page = Page()
        endpage = Page()
        pdf.append_page(page)
        pdf.append_page(endpage)

        page_layout = SingleColumnLayout(page)
        page_layout.vertical_margin = page.get_page_info().get_height() * Decimal(0.0001)

        endpage_layout = SingleColumnLayout(endpage)
        endpage_layout.vertical_margin = endpage.get_page_info().get_height() * Decimal(0.01)

        add_artwork_upper_right_corner(page)
        add_artwork_upper_right_corner(endpage)
        # add_gray_artwork_upper_right_corner(endpage)

        # Invoice information table
        page_layout.add(_build_invoice_information(df_inv))

        # Empty paragraph for spacing
        page_layout.add(Paragraph(" "))

        # Billing and shipping information table
        page_layout.add(_build_billing_and_shipping_information(name))

        # Itemized description
        page_layout.add(_build_itemized_description_table(df_inv,p_amt, p_out, p_date))

        add_colored_artwork_bottom_right_corner(page)
        add_colored_artwork_bottom_right_corner(endpage)

        # Outline
        # pdf.add_outline("Your Invoice", 0, DestinationType.FIT, page_nr=0)
        filename = 'Receipt\\' + receipt_no +".pdf"
        pdf_path = os.path.join(current_path, filename)
        with open(pdf_path, "wb") as pdf_file_handle:
            PDF.dumps(pdf_file_handle, pdf)

        for row in inv_sheet.iter_rows():
            if row[1].value == p_inv:
                row[7].value = ttl_paid
                row[11].value = p_out
        inv_sheet.cell(column=17, row=1, value=last_rno + 1)
        wb.save(filepath)
        wb.close()
        st.write("Receipts/s are created successfully.")
        st.experimental_rerun()

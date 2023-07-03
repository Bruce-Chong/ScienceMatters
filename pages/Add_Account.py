
import streamlit as st
import pandas as pd
import os
from pathlib import Path
import openpyxl
from datetime import datetime

st.set_page_config(page_title='Science Matters')
st.title('Science Matters')
st.subheader('Choose an action')


##################retrieve info from excel ***************************

# Path Setting:
try:
    current_path = Path(__file__).parent.absolute()  # Get the file path for this py file
except:
    current_path = (Path.cwd())

# import the inv_sheet from the excel file

filepath = os.path.join(current_path, 'ERP.xlsx')
df = openpyxl.load_workbook(filepath)
#df = pd.read_excel(filepath, engine='openpyxl')
psheet = df['Parent']
arsheet = df['AR']
prow_count = psheet.max_row
arrow_count = arsheet.max_row
last_acct_no = psheet.cell(row=2, column=11).value

###################create GUI##################

with st.form("new_acc"):
    st.write("Enter all details")
    pname = st.text_input("Parent name")
    phone = st.text_input("Telephone number")
    email = st.text_input("Email")
    sname = st.text_input("Student name")
    level = st.selectbox("Primary level", ('P3', 'P4', 'P5', 'P6'))
    day = st.selectbox("Class", ('1A', '1B', '2A', '2B', '3A', '3B', '4A', '4B', '5A', '5B'))
    # Every form must have a submit button.
    submitted = st.form_submit_button("Submit")
    if submitted:
        psheet.cell(column=1, row=prow_count + 1, value=last_acct_no)
        psheet.cell(column=2, row=prow_count + 1, value=pname)
        psheet.cell(column=4, row=prow_count + 1, value=phone)
        psheet.cell(column=5, row=prow_count + 1, value=email)
        psheet.cell(column=8, row=prow_count + 1, value=datetime.today().strftime('%d/%m/%Y'))
        psheet.cell(column=3, row=prow_count + 1, value=sname)
        psheet.cell(column=6, row=prow_count + 1, value=level)
        psheet.cell(column=7, row=prow_count + 1, value=day)

        arsheet.cell(column=1, row=arrow_count + 1, value=last_acct_no)
        arsheet.cell(column=2, row=arrow_count + 1, value='0')
        arsheet.cell(column=3, row=arrow_count + 1, value='SGD')
        arsheet.cell(column=5, row=arrow_count + 1, value='0')
        arsheet.cell(column=6, row=arrow_count + 1, value='0')
        arsheet.cell(column=7, row=arrow_count + 1, value='0')

        psheet.cell(row=2, column=11, value=last_acct_no + 1)
        df.save(filepath)
        df.close()
        st.write("New account " + str(last_acct_no) + "successfully created!")
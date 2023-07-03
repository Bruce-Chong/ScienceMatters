
import streamlit as st
import pandas as pd
import os
from pathlib import Path
import openpyxl
from datetime import date
#####################pdf imports###################
from borb.pdf import Document
from borb.pdf.page.page import Page
from borb.pdf.canvas.layout.page_layout.multi_column_layout import SingleColumnLayout
from decimal import Decimal
from borb.pdf.canvas.layout.image.image import Image
from borb.pdf.canvas.layout.table.fixed_column_width_table import FixedColumnWidthTable as Table
from borb.pdf.canvas.layout.text.paragraph import Paragraph
from borb.pdf.canvas.layout.layout_element import Alignment
from borb.pdf.pdf import PDF
from borb.pdf.canvas.color.color import HexColor, X11Color
from borb.pdf.canvas.layout.table.fixed_column_width_table import FixedColumnWidthTable as Table
from borb.pdf.canvas.layout.table.table import TableCell
from borb.pdf.canvas.layout.image.barcode import Barcode, BarcodeType
from borb.pdf.canvas.layout.layout_element import LayoutElement
from borb.pdf.canvas.line_art.line_art_factory import LineArtFactory
from borb.pdf.canvas.geometry.rectangle import Rectangle
from borb.pdf.canvas.layout.shape.shape import Shape
from borb.pdf.page.page_size import PageSize
import typing
import random

############################paynowQR#######################
from pyPayNowSg import PayNowConfigFactory, PayNowSerializer


st.set_page_config(page_title='Science Matters')
st.title('Invoice')

##################retrieve info from excel ***************************

# Path Setting:
try:
    current_path = Path(__file__).parent.absolute()  # Get the file path for this py file
except:
    current_path = (Path.cwd())

filepath = os.path.join(current_path, 'ERP.xlsx')
wb = openpyxl.load_workbook(filepath)
#df = pd.read_excel(filepath, engine='openpyxl')
inv_sheet = wb['Invoice']
mat_sheet = wb['Material']
ar_sheet = wb['AR']


inv_maxrow = inv_sheet.max_row
if 'ttl_amt' not in st.session_state:
    st.session_state.ttl_amt = 0

################methods###############################
def qr_payload(pay_ref, fee):
    # 0 for phone, 2 for UEN
    merchant_info = PayNowConfigFactory.build_merchant_account_info(
        2,
        "202238718D",
        True
    )

    reference = PayNowConfigFactory.build_additional_data(pay_ref)
    qr_pl = PayNowSerializer.serialize("Jane Chong", merchant_info, fee, reference)
    return qr_pl

def add_artwork_upper_right_corner(page: Page) -> None:
    ps: typing.Tuple[Decimal, Decimal] = PageSize.A4_PORTRAIT.value
    N: int = 4
    M: Decimal = Decimal(32)
    x: Decimal = ps[0] - N * 105
    y: Decimal = ps[1] - (0 + 2) * M
    IMAGE_PATH = Path(r"C:\Users\Choon Yong Chong\OneDrive\Documents\Business\Tuition center\SM logo.png")
    Image(image=IMAGE_PATH, width=239, height=27).layout(page, Rectangle(x, y, M, M))
    #page_layout.add(Image(image=IMAGE_PATH, width=100, height=50, margin_top=0, margin_right=0))  # change the size as you wish


def add_gray_artwork_upper_right_corner(page: Page) -> None:
    """
    This method will add a gray artwork of squares and triangles in the upper right corner
    of the given Page
    """
    grays: typing.List[HexColor] = [
        HexColor("A9A9A9"),
        HexColor("D3D3D3"),
        HexColor("DCDCDC"),
        HexColor("E0E0E0"),
        HexColor("E8E8E8"),
        HexColor("F0F0F0"),
    ]
    ps: typing.Tuple[Decimal, Decimal] = PageSize.A4_PORTRAIT.value
    N: int = 4
    M: Decimal = Decimal(32)

    # Draw triangles
    for i in range(0, N):
        x: Decimal = ps[0] - N * M + i * M
        y: Decimal = ps[1] - (i + 1) * M
        rg: HexColor = random.choice(grays)
        Shape(
            points=[(x + M, y), (x + M, y + M), (x, y + M)],
            stroke_color=rg,
            fill_color=rg,
        ).layout(page, Rectangle(x, y, M, M))

    # Draw squares
    for i in range(0, N - 1):
        for j in range(0, N - 1):
            if j > i:
                continue
            x: Decimal = ps[0] - (N - 1) * M + i * M
            y: Decimal = ps[1] - (j + 1) * M
            rg: HexColor = random.choice(grays)
            Shape(
                points=[(x, y), (x + M, y), (x + M, y + M), (x, y + M)],
                stroke_color=rg,
                fill_color=rg,
            ).layout(page, Rectangle(x, y, M, M))


def add_colored_artwork_bottom_right_corner(page: Page) -> None:
    """
    This method will add a blue/purple artwork of lines
    and squares to the bottom right corner
    of the given Page
    """
    ps: typing.Tuple[Decimal, Decimal] = PageSize.A4_PORTRAIT.value
    # Line
    r: Rectangle = Rectangle(Decimal(0), Decimal(32), ps[0], Decimal(8))
    Shape(
        points=LineArtFactory.rectangle(r),
        stroke_color=HexColor("fbb03f"),
        fill_color=HexColor("fbb03f"),
    ).layout(page, r)

def _build_invoice_information(inv, dtfrom, dtto, duedt):

    table_001 = Table(number_of_rows=5, number_of_columns=6)

    table_001.add(TableCell(Paragraph("Science Matters Pte Ltd", font="Helvetica-Bold"), col_span=2))
    table_001.add(TableCell(Paragraph(" "), col_span=2))
    table_001.add(TableCell(Paragraph("INVOICE", font_color=HexColor("#fbb03f"), font_size=Decimal(14),font="Helvetica-Bold", horizontal_alignment=Alignment.CENTERED), col_span=2))

    table_001.add(TableCell(Paragraph("https://www.sciencematters.com.sg", font_size=Decimal(10)), col_span=2))
    table_001.add(TableCell(Paragraph(" "), col_span=4))

    if dtfrom == "None":
        dt_frm_desp = " "
        dtfrom = " "
    else:
        dt_frm_desp = "Bill Date From"
        dtfrom = ": %d/%d/%d" % (dtfrom.day, dtfrom.month, dtfrom.year)

    if dtto == "None":
        dt_to_desp = " "
        dtto = " "
    else:
        dt_to_desp = "Bill Date From"
        dtto = ": %d/%d/%d" % (dtto.day, dtto.month, dtto.year)

    table_001.add(Paragraph(dt_frm_desp, font_size=Decimal(10), font="Helvetica-Bold"))
    table_001.add(Paragraph(dtfrom, font_size=Decimal(10)))
    table_001.add(TableCell(Paragraph(" "), col_span=2))
    table_001.add(Paragraph("Date :", font_size=Decimal(10), font="Helvetica-Bold", horizontal_alignment=Alignment.RIGHT))
    now = date.today().strftime('%d/%m/%Y')
    table_001.add(Paragraph(now, font_size=Decimal(10)))


    table_001.add(Paragraph(dt_to_desp, font_size=Decimal(10), font="Helvetica-Bold"))
    table_001.add(Paragraph(dtto, font_size=Decimal(10)))
    table_001.add(TableCell(Paragraph(" "), col_span=2))
    table_001.add(Paragraph("Invoice :", font_size=Decimal(10), font="Helvetica-Bold", horizontal_alignment=Alignment.RIGHT))
    table_001.add(Paragraph(inv, font_size=Decimal(10)))

    table_001.add(TableCell(Paragraph(" "), col_span=4))
    table_001.add(Paragraph("Due Date :", font_size=Decimal(10), font="Helvetica-Bold", horizontal_alignment=Alignment.RIGHT))
    table_001.add(Paragraph("%d/%d/%d" % (duedt.day, duedt.month, duedt.year), font_size=Decimal(10)))


    table_001.set_padding_on_all_cells(Decimal(2), Decimal(2), Decimal(2), Decimal(2))
    table_001.no_borders()
    return table_001


def _build_billing_and_shipping_information():
    table_001 = Table(number_of_rows=3, number_of_columns=1)
    table_001.add(Paragraph("BILL TO",font_size=Decimal(10), font="Helvetica-Bold"))
    table_001.add(Paragraph(name,font_size=Decimal(10)))  # BILLING
    table_001.add(Paragraph(" "))
    table_001.set_padding_on_all_cells(Decimal(2), Decimal(2), Decimal(2), Decimal(2))
    table_001.no_borders()
    return table_001


def _build_itemized_description_table(m_acct, m_name):
    #df_inv = pd.DataFrame(columns=['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty', 'Price'])
    df_inv = st.session_state.bill.loc[(st.session_state.bill['Account'] == int(m_acct)) & (st.session_state.bill['Student'] == m_name)]
    #df_inv = st.session_state.bill.loc[st.session_state.bill['Account'] == int(m_acct)]
    mat_df = pd.read_excel(filepath, engine='openpyxl',sheet_name='Material')
    ar_df = pd.read_excel(filepath, engine='openpyxl',sheet_name='AR')
    #count number of items/acct code appears in dataframe
    #nrow = df_inv.shape[0] + 7
    nrow = df_inv.shape[0] + 5
    st.write('nrow is' +str(nrow) + ' account is ' + str(m_acct) + ' name is ' + m_name)
    # draw table headers
    table_001 = Table(number_of_rows=nrow, number_of_columns=6)

    table_001.add(TableCell(Paragraph("Description", font_size=Decimal(9)), background_color=HexColor("#D3D3D3"),
                            padding_bottom=Decimal(30),
                            col_span=3))
    title = ["Qty", "UNIT PRICE", "AMOUNT"]
    for n in range(0, 3):
        table_001.add(TableCell(Paragraph(title[n],
                                          font_size=Decimal(9),
                                          horizontal_alignment=Alignment.RIGHT),
                                background_color=HexColor("#D3D3D3"),
                                padding_bottom=Decimal(30), ))

    # retrieve all values from excel list within the list
    sub_amount = 0
    curr_amount = 0
    for row_dict in df_inv.to_dict(orient="records"):
        # Item description
        plvl = row_dict['Item']
        unit_pr = "{:.2f}".format(float(row_dict['Price']))
        itm_ttpr = "{:.2f}".format(float(row_dict['Price']) * float(row_dict['Qty']))
        qty = str(row_dict['Qty'])
        try:
            description = mat_df.loc[mat_df['mcode'] == plvl, 'desc'].iloc[0]
        except:
            description = plvl
        table_001.add(TableCell(Paragraph(description, font_size=Decimal(9)), col_span=3))
        table_001.add(TableCell(Paragraph(qty, font_size=Decimal(9), horizontal_alignment=Alignment.RIGHT)))
        table_001.add(TableCell(Paragraph("$ " + str(unit_pr), font_size=Decimal(9), horizontal_alignment=Alignment.RIGHT)))
        table_001.add(TableCell(Paragraph("$ " + str(itm_ttpr), font_size=Decimal(9), horizontal_alignment=Alignment.RIGHT)))
        sub_amount = "{:.2f}".format(float(row_dict['Price']) * float(row_dict['Qty']) + float(sub_amount))
    #calculate all total values here

    # Reading values from account sheet
    prev_chrg = float(ar_df.loc[ar_df['acct_no'] == int(m_acct), 'pay_due'].iloc[0])
    paid = float(ar_df.loc[ar_df['acct_no'] == int(m_acct), 'prev_paymt'].iloc[0])
    #total_amount = "{:.2f}".format(float(sub_amount) + float(prev_chrg) - float(paid))
    total_amount = sub_amount
    st.session_state.ttl_amt = total_amount
    curr_amount = sub_amount
    #paid = str(paid)
    #prev_chrg = str(prev_chrg)

    table_001.add(TableCell(Paragraph("_____________________________________________________________________",font_color=HexColor("#D3D3D3"), horizontal_alignment=Alignment.CENTERED),col_span=6))
    table_001.add(TableCell(Paragraph("SubTotal :",font_size=Decimal(9) , horizontal_alignment=Alignment.RIGHT), col_span=5))
    table_001.add(TableCell(Paragraph("$ " + str(sub_amount),font_size=Decimal(9), horizontal_alignment=Alignment.RIGHT)))
    table_001.add(
        TableCell(Paragraph("Current Charges :", font_size=Decimal(9), horizontal_alignment=Alignment.RIGHT), background_color=HexColor("#D3D3D3"), col_span=5))
    table_001.add(TableCell(Paragraph("$ " + str(curr_amount),font_size=Decimal(9), horizontal_alignment=Alignment.RIGHT), background_color=HexColor("#D3D3D3")))

    #table_001.add(TableCell(Paragraph("Previous Balance :", font_size=Decimal(9), horizontal_alignment=Alignment.RIGHT), col_span=5))
    #table_001.add(TableCell(Paragraph("$ " + prev_chrg, font_size=Decimal(9), horizontal_alignment=Alignment.RIGHT)))

    #table_001.add(TableCell(Paragraph("Payment received :", font_size=Decimal(9),horizontal_alignment=Alignment.RIGHT), col_span=5))
    #table_001.add(TableCell(Paragraph("-$ " + paid, font_size=Decimal(9), horizontal_alignment=Alignment.RIGHT), padding_bottom=Decimal(30)))

    table_001.add(TableCell(Paragraph("Outstanding Amount Due on  " + "%d/%d/%d" % (due_dt.day, due_dt.month, due_dt.year) + " :",
                            font_size=Decimal(9),
                            font="Helvetica-Bold",
                            horizontal_alignment=Alignment.RIGHT),
                            background_color=HexColor("#fbb03f"),
                            col_span=5))

    table_001.add(TableCell(Paragraph("$ " + str(total_amount),
                                      font_size=Decimal(9),
                                      font="Helvetica-Bold",
                                      horizontal_alignment=Alignment.RIGHT),
                            background_color=HexColor("#fbb03f")))

    table_001.set_padding_on_all_cells(Decimal(2), Decimal(2), Decimal(2), Decimal(2))
    table_001.no_borders()
    return table_001

def bottom_remarks():
    table_001 = Table(number_of_rows=10, number_of_columns=1)
    table_001.add(Paragraph("Please kindly make payment by " + "%d/%d/%d" % (due_dt.day, due_dt.month, due_dt.year) +". Payment can be made through PayNow (preferred) or direct bank transfer. When making payment, please include your invoice number in the reference field.",
                            font_size=Decimal(9),
                            font="Helvetica-Oblique",
                            padding_bottom=Decimal(20)
                            ))

    table_001.add(Paragraph("For PayNow",
                            font_size=Decimal(9),
                            font="Helvetica-Bold",
                            ))
    table_001.add(Paragraph("UEN : 202238718D",
                            font_size=Decimal(9),
                            ))

    qr_code: LayoutElement = Barcode(
        data=qr_payload(invoice_number, st.session_state.ttl_amt),
        width=Decimal(64),
        height=Decimal(64),
        type=BarcodeType.QR,
     )

    table_001.add(qr_code)

    table_001.add(Paragraph("Payments by fund transfer should be made to the following account",
                            font_size=Decimal(9),
                            font="Helvetica-Bold",
                            ))
    table_001.add(Paragraph("Bank Name: United Overseas Bank Limited Singapore",
                            font_size=Decimal(9),
                            ))
    table_001.add(Paragraph("Bank Account Name: Science Matters Pte Ltd",
                            font_size=Decimal(9),
                            ))
    table_001.add(Paragraph("Bank Account Number: 7693418236",
                            font_size=Decimal(9),
                            ))
    table_001.add(Paragraph("If my lessons have benefitted your child, why not refer a friend to me today? Your child will receive a complimentary lesson for each new enrolment of a student you have directly referred.",
                            font_size=Decimal(9),
                            horizontal_alignment=Alignment.CENTERED,
                            padding_top=Decimal(10)
                            ))

    table_001.add(Paragraph("Thank you for your support!",
                            font_size=Decimal(12),
                            font="Helvetica-Bold",
                            horizontal_alignment=Alignment.CENTERED,
                            ))

    table_001.set_padding_on_all_cells(Decimal(2), Decimal(2), Decimal(2), Decimal(2))
    table_001.no_borders()
    return table_001

def terms_conditions():
    table_001 = Table(number_of_rows=14, number_of_columns=10)
    padding = Decimal(5)
    sec_pad = Decimal(15)


    table_001.add(TableCell(Paragraph("Terms and Conditions for 2023",
        font_size=Decimal(12),
        font="Helvetica-Bold",
        horizontal_alignment=Alignment.CENTERED
        ),
        col_span = 10))

    table_001.add(Paragraph("A.", font_size=Decimal(10), font="Helvetica-Bold", padding_top=sec_pad))
    table_001.add(TableCell(Paragraph("PAYMENT METHOD", font_size=Decimal(10),padding_top=sec_pad, font="Helvetica-Bold"), col_span=9))

    table_001.add(Paragraph("1.", font_size=Decimal(10), padding_top=padding))
    table_001.add(TableCell(Paragraph("Please use PayNow or direct bank transfer as the payment method.",
                                      font_size=Decimal(10), padding_top = padding),
                            col_span=9))


    table_001.add(Paragraph("2.", font_size=Decimal(10), padding_top=padding))
    table_001.add(TableCell(Paragraph("Please include the invoice number in the reference field.",
        font_size=Decimal(10), padding_top=padding),
        col_span=9))

    #table_001.add(Paragraph("4.", font_size=Decimal(10), padding_top=padding))
    #table_001.add(TableCell(Paragraph("Fees for all the lessons of the month are due on the first day of each month.",
                                      #font_size=Decimal(10), padding_top=padding),
                            #col_span=9))

    table_001.add(Paragraph("B.", font_size=Decimal(10), font="Helvetica-Bold", padding_top=sec_pad))
    table_001.add(TableCell(Paragraph("LATE FEES", font_size=Decimal(10), padding_top=sec_pad, font="Helvetica-Bold"),
                  col_span=9))

    table_001.add(Paragraph("1.", font_size=Decimal(10), padding_top=padding))
    table_001.add(TableCell(Paragraph("To encourage prompt payment of fees and to defray administrative cost of fee reminders, a late fee of $50 will be charged for fees paid past the due date.",
                                      font_size=Decimal(10), padding_top=padding),
                            col_span=9))

    table_001.add(Paragraph("C.", font_size=Decimal(10), font="Helvetica-Bold", padding_top=sec_pad))
    table_001.add(
        TableCell(Paragraph("RESCHEDULING AND CANCELLATIONS OF CLASSES", font_size=Decimal(10), padding_top=sec_pad, font="Helvetica-Bold"), col_span=9))

    table_001.add(Paragraph("1.", font_size=Decimal(10), padding_top=padding))
    table_001.add(TableCell(Paragraph(
        "Unless otherwise stated, lessons are carried out weekly. Regular weekly attendance is expected during school terms to ensure that students have consistency in their learning and revision.",
        font_size=Decimal(10), padding_top=padding), col_span=9))

    table_001.add(Paragraph("2.", font_size=Decimal(10), padding_top=padding))
    table_001.add(TableCell(Paragraph("There are no refunds for missed lessons. With a minimum of 24 hours notice, the following alternatives may be offered:",
        font_size=Decimal(10), padding_top=padding),
                            col_span=9))

    table_001.add(Paragraph(" ", font_size=Decimal(10), padding_top=padding))
    table_001.add(Paragraph("a)",font_size=Decimal(10), padding_top=padding))
    table_001.add(TableCell(Paragraph(" to join another class at a different timing (subject to availability)",
       font_size=Decimal(10), padding_top=padding),
                            col_span=8))

    table_001.add(Paragraph(" ", font_size=Decimal(10), padding_top=padding))
    table_001.add(Paragraph("b)", font_size=Decimal(10), padding_top=padding))
    table_001.add(TableCell(Paragraph("to receive a zoom recording of the lesson", font_size=Decimal(10), padding_top=padding), col_span=8))

    table_001.add(Paragraph("3.", font_size=Decimal(10), padding_top=padding))
    table_001.add(TableCell(Paragraph("The fee for the missed lesson may be credited to the next month only with a valid medical certificate for the day of the lesson.",
        font_size=Decimal(10), padding_top=padding), col_span=9))



    table_001.add(Paragraph("D.", font_size=Decimal(10), font="Helvetica-Bold", padding_top=sec_pad))
    table_001.add(TableCell(Paragraph("OTHER FEES", font_size=Decimal(10), padding_top=sec_pad,
                            font="Helvetica-Bold"), col_span=9))

    table_001.add(Paragraph("1.", font_size=Decimal(10), padding_top=padding))
    table_001.add(TableCell(Paragraph("Books and other lesson materials are charged separately from tuition fees.",
        font_size=Decimal(10), padding_top=padding),
        col_span=9))

    table_001.set_padding_on_all_cells(Decimal(2), Decimal(2), Decimal(2), Decimal(2))
    table_001.no_borders()
    return table_001

###################create options as buttons##################
with st.expander('Invoice billing due list'):
    st.session_state.bill.sort_values(by=['Class', 'Account'], inplace=True)
    st.dataframe(st.session_state.bill)
    last_invno = int(inv_sheet.cell(row=1, column=16).value)
    if st.button("Invoice all in list"):
        x_acct = ''
        sname = ''
        for row_dict in st.session_state.bill.to_dict(orient="records"):
            #set the first account code as the first item
            if x_acct != str(row_dict['Account']) or ((row_dict['Account'] == 199999 and sname != str(row_dict['Student']))):
                x_acct = str(row_dict['Account'])
                invoice_number = str(row_dict['Account']) + str(last_invno)
                last_invno = last_invno + 1
                invoice_dt = str(date.today().strftime('%d/%m/%Y'))
                #invoice_frm = st.session_state.dtfrm
                #due_dt = st.session_state.duedt
                #invoice_dtto = st.session_state.dtto
                due_dt = row_dict['Due_Date']
                invoice_dtto = row_dict['Date_To']
                invoice_frm = row_dict['Date_From']
                pname = row_dict['Parent']
                sname = row_dict['Student']
                if sname == "None":
                    name = pname
                else:
                    name = f'{pname}, for student {sname}'

                # Create document
                pdf = Document()

                #calculate no of items in invoice
                df_count = st.session_state.bill.loc[(st.session_state.bill['Account'] == int(x_acct)) & (st.session_state.bill['Student'] == sname)]
                no_itm = df_count.shape[0]
                # Add page
                page = Page()
                page2 = Page()
                endpage = Page()

                pdf.append_page(page)
                if no_itm > 4:
                    pdf.append_page(page2)
                pdf.append_page(endpage)

                page_layout = SingleColumnLayout(page)
                page_layout.vertical_margin = page.get_page_info().get_height() * Decimal(0.0001)

                page2_layout = SingleColumnLayout(page2)
                page2_layout.vertical_margin = page2.get_page_info().get_height() * Decimal(0.0001)

                endpage_layout = SingleColumnLayout(endpage)
                endpage_layout.vertical_margin = endpage.get_page_info().get_height() * Decimal(0.01)

                add_artwork_upper_right_corner(page)
                add_artwork_upper_right_corner(page2)
                add_artwork_upper_right_corner(endpage)
                # add_gray_artwork_upper_right_corner(endpage)

                # Invoice information table
                page_layout.add(_build_invoice_information(invoice_number, invoice_frm, invoice_dtto, due_dt))

                # Empty paragraph for spacing
                #page_layout.add(Paragraph(" "))

                # Billing and shipping information table
                page_layout.add(_build_billing_and_shipping_information())

                # Itemized description
                page_layout.add(_build_itemized_description_table(x_acct, sname))

                if no_itm > 4:
                    page2_layout.add(bottom_remarks())
                else:
                    page_layout.add(bottom_remarks())

                # terms and conditions
                endpage_layout.add(terms_conditions())

                add_colored_artwork_bottom_right_corner(page)
                add_colored_artwork_bottom_right_corner(page2)
                add_colored_artwork_bottom_right_corner(endpage)

                # Outline
                # pdf.add_outline("Your Invoice", 0, DestinationType.FIT, page_nr=0)
                filename = 'Invoice pdf\\' + invoice_number + ".pdf"
                pdf_path = os.path.join(current_path, filename)
                with open(pdf_path, "wb") as pdf_file_handle:
                    PDF.dumps(pdf_file_handle, pdf)

                #add invoice header into excel
                inv_sheet.cell(column=1, row=inv_maxrow + 1, value=x_acct)
                inv_sheet.cell(column=2, row=inv_maxrow + 1, value=invoice_number)
                inv_sheet.cell(column=3, row=inv_maxrow + 1, value='SGD')
                inv_sheet.cell(column=4, row=inv_maxrow + 1, value=st.session_state.ttl_amt)
                if invoice_frm == 'None':
                    inv_sheet.cell(column=5, row=inv_maxrow + 1, value=invoice_frm)
                else:
                    inv_sheet.cell(column=5, row=inv_maxrow + 1, value=invoice_frm.strftime('%d/%m/%Y'))
                if invoice_dtto == 'None':
                    inv_sheet.cell(column=6, row=inv_maxrow + 1, value=invoice_dtto)
                else:
                    inv_sheet.cell(column=6, row=inv_maxrow + 1, value=invoice_dtto.strftime('%d/%m/%Y'))
                inv_sheet.cell(column=7, row=inv_maxrow + 1, value=date.today().strftime('%d/%m/%Y'))
                inv_sheet.cell(column=9, row=inv_maxrow + 1, value='no')
                inv_sheet.cell(column=10, row=inv_maxrow + 1, value=due_dt.strftime('%d/%m/%Y'))
                inv_sheet.cell(column=11, row=inv_maxrow + 1, value='no')
                inv_sheet.cell(column=12, row=inv_maxrow + 1, value=st.session_state.ttl_amt)
                inv_sheet.cell(column=13, row=inv_maxrow + 1, value=name)

                inv_maxrow = inv_maxrow + 1

                for row in ar_sheet.iter_rows():
                    if row[0].value == int(x_acct) and x_acct !="199999":
                        row[1].value = float(row[1].value) + float(st.session_state.ttl_amt)
                        row[4].value = float(st.session_state.ttl_amt)
                        row[5].value = 0
                        row[6].value = invoice_number

        inv_sheet.cell(column=16, row=1, value=last_invno)

        wb.save(filepath)
        wb.close()
        st.write("Invoice/s are created successfully.")

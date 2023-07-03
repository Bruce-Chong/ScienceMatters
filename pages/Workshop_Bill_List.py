
import streamlit as st
import pandas as pd
import os
from pathlib import Path
import openpyxl
import re
from datetime import date
import numpy as np
from streamlit_extras.switch_page_button import switch_page
from datetime import timedelta

st.set_page_config(page_title='Science Matters')
st.title('Invoice')
##################retrieve info from excel ***************************

# Path Setting:
try:
    current_path = Path(__file__).parent.absolute()  # Get the file path for this py file
except:
    current_path = (Path.cwd())

# import the inv_sheet from the excel file

filepath = os.path.join(current_path, 'ERP.xlsx')
wb = openpyxl.load_workbook(filepath)
#df = pd.read_excel(filepath, engine='openpyxl')
wssheet = wb['Workshop']

if 'bill' not in st.session_state:
    st.session_state['bill'] = pd.DataFrame(columns = ['Account', 'Parent', 'Student', 'Class', 'Item', 'Qty','Price', 'Date_From', 'Date_To', 'Due_Date'])

def add_one_time_billing(xname, s_name, xitm, xqty, xprice):
    add_row = pd.DataFrame({'Account': 199999, 'Parent': xname, 'Student': s_name, 'Class': "None", 'Item': xitm, 'Qty': xqty, 'Price': xprice, 'Date_From': "None", 'Date_To': "None", 'Due_Date': [st.session_state.duedt]})
    st.session_state.bill = pd.concat([st.session_state.bill, add_row], ignore_index = True)

############!!!Cockpit: Set column values here, very important!!!############
ws_qty = 5  #no of workshops
pname_no = 2  #parent name. column no starts from 0
sname_no = 3
email_no = 4
ph_no = 5
lvl_no = 6
ws_no  = 7
ws_ttl = ws_qty + ws_no - 1  #to calculate index of the last column for workshop
dis_no = ws_qty + ws_no  #get the last column as discount
wsname = []

#put workshop description in header into a list
for row in wssheet.iter_cols(min_row=1, min_col=ws_no+1, max_row=1, max_col=dis_no):
    for cell in row:
        if cell.value is not None:
            tmp_str = re.split('\(|\$|\)', cell.value)
            wsname.append(tmp_str)
            #st.write(tmp_str)


wssheet.delete_rows(1)   #to remove header information

st.subheader('Billing workshop here')
with st.form(key = "workshop"):
    c1, c2, c3 = st.columns(3)
    with c1:
        inv_duedt = st.date_input("Invoice Due Date")

    if st.form_submit_button(label= 'workshop bill'):

        st.session_state.duedt = inv_duedt
        st.session_state.acct = "199999"
        for row in wssheet.iter_rows():
            pname = row[pname_no].value
            sname = row[sname_no].value
            email = row[email_no].value
            ph = row[sname_no].value
            lvl = row[lvl_no].value
            for x in range(ws_no, ws_ttl+1):
                wssession = row[x].value
                if wssession is not None:
                    wssession = wsname[x-ws_no]
                    add_one_time_billing(pname, sname, wssession[0] + row[x].value, 1, wssession[2])
            discount = row[dis_no].value
            if discount != 0:
                add_one_time_billing(pname, sname, 'Discount 5%', 1, discount)
        st.experimental_rerun()

# display all billing items
st.session_state.bill.sort_values(by=['Class', 'Account'], inplace=True)
st.dataframe(st.session_state.bill)







